"""Import en masse d'établissements depuis un fichier Excel (.xlsx).

Flux en deux temps, volontairement sans écriture partielle en base :
1. `valider_fichier_excel` — parse le fichier et vérifie chaque ligne
   (formats, cohérence du découpage territorial, doublons) sans jamais
   toucher la base. Retourne un rapport complet, ligne par ligne.
2. `confirmer_creation` (dans les routes) — reçoit les lignes validées et
   choisies par l'utilisateur, les revalide (défense en profondeur) et les
   insère en base.
"""
import re
import unicodedata
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from motor.motor_asyncio import AsyncIOMotorDatabase

from app.models_sql import Departement, Commune, Arrondissement, Quartier
from app.schemas.etablissement import EtablissementCreate
from app.schemas.etablissement_import import LigneImportValide, LigneImportInvalide, RapportValidationImport
from app.validators import normaliser_telephone_benin

TYPES_ETABLISSEMENT_VALIDES = ["CHU", "CHD", "CSC", "Clinique", "Maternite"]
STATUTS_ETABLISSEMENT_VALIDES = ["actif", "maintenance", "inactif"]

# Bornes géographiques approximatives du Bénin (avec marge) — détecte les
# coordonnées aberrantes ou inversées (latitude/longitude interverties).
LATITUDE_MIN_BENIN, LATITUDE_MAX_BENIN = 5.5, 13.0
LONGITUDE_MIN_BENIN, LONGITUDE_MAX_BENIN = 0.3, 4.3

COLONNES_ATTENDUES = [
    "nom", "departement", "commune", "arrondissement", "quartier",
    "adresse", "type", "statut", "telephone", "latitude", "longitude",
]
COLONNES_REQUISES = [
    "nom", "departement", "commune", "arrondissement", "quartier",
    "type", "telephone", "latitude", "longitude",
]


def _normaliser_texte(v: str | None) -> str:
    """Supprime accents, espaces superflus et met en minuscule, pour une
    comparaison tolérante aux variations de saisie (ex: 'Département' /
    'departement', 'Cotonou ' / 'cotonou')."""
    if v is None:
        return ""
    sans_accents = unicodedata.normalize("NFKD", str(v)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", sans_accents.strip().lower())


def _valeur_cellule(cellule) -> str | None:
    if cellule is None:
        return None
    if isinstance(cellule, float) and cellule.is_integer():
        return str(int(cellule))
    v = str(cellule).strip()
    return v or None


class ReferentielTerritoire:
    """Chargé une fois par validation, tient tout le découpage territorial
    en mémoire (12 départements, 77 communes, 546 arrondissements, ~5300
    quartiers — volumes négligeables) pour résoudre chaque ligne sans
    aller-retour base par ligne."""

    def __init__(self):
        self._departements: dict[str, dict] = {}

    @classmethod
    async def charger(cls, db_sql: AsyncSession) -> "ReferentielTerritoire":
        ref = cls()
        deps = (await db_sql.execute(select(Departement))).scalars().all()
        coms = (await db_sql.execute(select(Commune))).scalars().all()
        arronds = (await db_sql.execute(select(Arrondissement))).scalars().all()
        quarts = (await db_sql.execute(select(Quartier))).scalars().all()

        deps_par_id = {d.id_dep: d for d in deps}
        coms_par_id = {c.id_com: c for c in coms}
        arronds_par_id = {a.id_arrond: a for a in arronds}

        for d in deps:
            ref._departements[_normaliser_texte(d.lib_dep)] = {"nom": d.lib_dep, "communes": {}}

        for c in coms:
            d = deps_par_id.get(c.id_dep)
            if not d:
                continue
            noeud_dep = ref._departements.get(_normaliser_texte(d.lib_dep))
            if noeud_dep is None:
                continue
            noeud_dep["communes"][_normaliser_texte(c.lib_com)] = {"nom": c.lib_com, "arrondissements": {}}

        for a in arronds:
            c = coms_par_id.get(a.id_com)
            if not c:
                continue
            d = deps_par_id.get(c.id_dep)
            if not d:
                continue
            noeud_com = ref._departements.get(_normaliser_texte(d.lib_dep), {}).get("communes", {}).get(_normaliser_texte(c.lib_com))
            if noeud_com is None:
                continue
            noeud_com["arrondissements"][_normaliser_texte(a.lib_arrond)] = {"nom": a.lib_arrond, "quartiers": {}}

        for q in quarts:
            a = arronds_par_id.get(q.id_arrond)
            if not a:
                continue
            c = coms_par_id.get(a.id_com)
            if not c:
                continue
            d = deps_par_id.get(c.id_dep)
            if not d:
                continue
            noeud_arrond = (
                ref._departements.get(_normaliser_texte(d.lib_dep), {})
                .get("communes", {}).get(_normaliser_texte(c.lib_com), {})
                .get("arrondissements", {}).get(_normaliser_texte(a.lib_arrond))
            )
            if noeud_arrond is None:
                continue
            noeud_arrond["quartiers"][_normaliser_texte(q.lib_quart)] = q.lib_quart

        return ref

    def resoudre(self, departement: str | None, commune: str | None, arrondissement: str | None, quartier: str | None):
        """Retourne (departement, commune, arrondissement, quartier) avec
        l'orthographe canonique du référentiel si la chaîne est valide,
        sinon (None, erreurs)."""
        erreurs: list[str] = []

        noeud_dep = self._departements.get(_normaliser_texte(departement))
        if noeud_dep is None:
            erreurs.append(f"Département inconnu : « {departement} ».")
            return None, erreurs

        noeud_com = noeud_dep["communes"].get(_normaliser_texte(commune))
        if noeud_com is None:
            erreurs.append(f"Commune « {commune} » introuvable dans le département {noeud_dep['nom']}.")
            return None, erreurs

        noeud_arrond = noeud_com["arrondissements"].get(_normaliser_texte(arrondissement))
        if noeud_arrond is None:
            erreurs.append(f"Arrondissement « {arrondissement} » introuvable dans la commune {noeud_com['nom']}.")
            return None, erreurs

        nom_quartier = noeud_arrond["quartiers"].get(_normaliser_texte(quartier))
        if nom_quartier is None:
            erreurs.append(f"Quartier « {quartier} » introuvable dans l'arrondissement {noeud_arrond['nom']}.")
            return None, erreurs

        return (noeud_dep["nom"], noeud_com["nom"], noeud_arrond["nom"], nom_quartier), []


def _lire_lignes_brutes(contenu: bytes) -> tuple[list[dict[str, str | None]], list[str]]:
    """Ouvre le classeur et associe chaque colonne à son en-tête normalisé.
    Retourne (lignes brutes indexées par nom de colonne, erreurs fatales)."""
    try:
        classeur = load_workbook(BytesIO(contenu), read_only=True, data_only=True)
    except Exception:
        return [], ["Le fichier n'est pas un classeur Excel (.xlsx) valide."]

    feuille = classeur.worksheets[0]
    lignes_iter = feuille.iter_rows(values_only=True)
    try:
        entetes_brutes = next(lignes_iter)
    except StopIteration:
        return [], ["Le fichier est vide."]

    index_colonnes: dict[str, int] = {}
    for i, entete in enumerate(entetes_brutes or []):
        nom_normalise = _normaliser_texte(entete)
        if nom_normalise in COLONNES_ATTENDUES:
            index_colonnes[nom_normalise] = i

    colonnes_manquantes = [c for c in COLONNES_REQUISES if c not in index_colonnes]
    if colonnes_manquantes:
        return [], [
            "Colonnes obligatoires manquantes dans l'en-tête : " + ", ".join(colonnes_manquantes) +
            ". Utilisez le modèle fourni pour garantir les bons intitulés de colonnes."
        ]

    lignes: list[dict[str, str | None]] = []
    for ligne_brute in lignes_iter:
        if ligne_brute is None or all(v is None or str(v).strip() == "" for v in ligne_brute):
            continue  # ligne entièrement vide, ignorée silencieusement
        valeurs = {
            col: _valeur_cellule(ligne_brute[i]) if i < len(ligne_brute) else None
            for col, i in index_colonnes.items()
        }
        lignes.append(valeurs)

    return lignes, []


def _valider_ligne(
    numero_ligne: int,
    valeurs: dict[str, str | None],
    referentiel: ReferentielTerritoire,
    telephones_deja_vus: dict[str, int],
) -> tuple[LigneImportValide | None, LigneImportInvalide | None]:
    erreurs: list[str] = []

    for champ in COLONNES_REQUISES:
        if not valeurs.get(champ):
            erreurs.append(f"Le champ « {champ} » est obligatoire.")

    type_valeur = valeurs.get("type")
    if type_valeur and type_valeur not in TYPES_ETABLISSEMENT_VALIDES:
        erreurs.append(f"Type invalide : « {type_valeur} ». Valeurs autorisées : {', '.join(TYPES_ETABLISSEMENT_VALIDES)}.")

    statut_valeur = valeurs.get("statut") or "actif"
    if statut_valeur not in STATUTS_ETABLISSEMENT_VALIDES:
        erreurs.append(f"Statut invalide : « {statut_valeur} ». Valeurs autorisées : {', '.join(STATUTS_ETABLISSEMENT_VALIDES)}.")

    telephone_normalise = None
    if valeurs.get("telephone"):
        try:
            telephone_normalise = normaliser_telephone_benin(valeurs["telephone"])
        except ValueError as e:
            erreurs.append(str(e))
        else:
            ligne_precedente = telephones_deja_vus.get(telephone_normalise)
            if ligne_precedente is not None:
                erreurs.append(f"Numéro de téléphone en doublon avec la ligne {ligne_precedente} du fichier.")
            else:
                telephones_deja_vus[telephone_normalise] = numero_ligne

    latitude = longitude = None
    for champ, cible in (("latitude", "latitude"), ("longitude", "longitude")):
        brut = valeurs.get(champ)
        if brut is None:
            continue
        try:
            valeur = float(str(brut).replace(",", "."))
        except ValueError:
            erreurs.append(f"« {champ} » doit être un nombre (reçu : « {brut} »).")
            continue
        if cible == "latitude":
            latitude = valeur
        else:
            longitude = valeur

    if latitude is not None and not (LATITUDE_MIN_BENIN <= latitude <= LATITUDE_MAX_BENIN):
        erreurs.append(f"Latitude {latitude} hors des limites du Bénin ({LATITUDE_MIN_BENIN} à {LATITUDE_MAX_BENIN}).")
    if longitude is not None and not (LONGITUDE_MIN_BENIN <= longitude <= LONGITUDE_MAX_BENIN):
        erreurs.append(f"Longitude {longitude} hors des limites du Bénin ({LONGITUDE_MIN_BENIN} à {LONGITUDE_MAX_BENIN}).")

    localisation = None
    if valeurs.get("departement") and valeurs.get("commune") and valeurs.get("arrondissement") and valeurs.get("quartier"):
        localisation, erreurs_localisation = referentiel.resoudre(
            valeurs.get("departement"), valeurs.get("commune"), valeurs.get("arrondissement"), valeurs.get("quartier")
        )
        erreurs.extend(erreurs_localisation)

    if erreurs:
        return None, LigneImportInvalide(numero_ligne=numero_ligne, valeurs_brutes=valeurs, erreurs=erreurs)

    departement, commune, arrondissement, quartier = localisation
    try:
        donnees = EtablissementCreate(
            nom=valeurs["nom"],
            departement=departement,
            commune=commune,
            arrondissement=arrondissement,
            quartier=quartier,
            adresse=valeurs.get("adresse"),
            latitude=latitude,
            longitude=longitude,
            type=type_valeur,
            statut=statut_valeur,
            telephone=telephone_normalise,
        )
    except ValidationError as e:
        return None, LigneImportInvalide(
            numero_ligne=numero_ligne, valeurs_brutes=valeurs,
            erreurs=[err["msg"] for err in e.errors()],
        )

    return LigneImportValide(numero_ligne=numero_ligne, donnees=donnees), None


async def verifier_doublons_telephone_base(
    db_mongo: AsyncIOMotorDatabase, telephones: list[str]
) -> dict[str, str]:
    """Retourne, pour chaque téléphone déjà utilisé en base (établissements
    ou prestataires), le nom de l'entité qui l'utilise."""
    if not telephones:
        return {}
    resultat: dict[str, str] = {}
    async for doc in db_mongo["etablissements"].find({"telephone": {"$in": telephones}}, {"telephone": 1, "nom": 1}):
        resultat[doc["telephone"]] = f"l'établissement « {doc['nom']} »"
    async for doc in db_mongo["prestataires_partenaires"].find({"telephone": {"$in": telephones}}, {"telephone": 1, "nom": 1}):
        resultat.setdefault(doc["telephone"], f"le prestataire « {doc['nom']} »")
    return resultat


async def valider_fichier_excel(
    contenu: bytes, db_sql: AsyncSession, db_mongo: AsyncIOMotorDatabase
) -> RapportValidationImport:
    lignes_brutes, erreurs_fatales = _lire_lignes_brutes(contenu)
    if erreurs_fatales:
        ligne_erreur = LigneImportInvalide(numero_ligne=0, valeurs_brutes={}, erreurs=erreurs_fatales)
        return RapportValidationImport(
            total_lignes=0, nombre_valides=0, nombre_invalides=1,
            lignes_valides=[], lignes_invalides=[ligne_erreur],
        )

    referentiel = await ReferentielTerritoire.charger(db_sql)

    lignes_valides: list[LigneImportValide] = []
    lignes_invalides: list[LigneImportInvalide] = []
    telephones_deja_vus: dict[str, int] = {}

    for i, valeurs in enumerate(lignes_brutes, start=1):
        valide, invalide = _valider_ligne(i, valeurs, referentiel, telephones_deja_vus)
        if valide:
            lignes_valides.append(valide)
        else:
            lignes_invalides.append(invalide)

    doublons_base = await verifier_doublons_telephone_base(
        db_mongo, [l.donnees.telephone for l in lignes_valides]
    )
    if doublons_base:
        encore_valides = []
        for ligne in lignes_valides:
            entite = doublons_base.get(ligne.donnees.telephone)
            if entite:
                valeurs_str = {k: (str(v) if v is not None else None) for k, v in ligne.donnees.model_dump().items()}
                lignes_invalides.append(LigneImportInvalide(
                    numero_ligne=ligne.numero_ligne,
                    valeurs_brutes=valeurs_str,
                    erreurs=[f"Le numéro {ligne.donnees.telephone} est déjà utilisé par {entite} en base."],
                ))
            else:
                encore_valides.append(ligne)
        lignes_valides = encore_valides

    lignes_invalides.sort(key=lambda l: l.numero_ligne)

    return RapportValidationImport(
        total_lignes=len(lignes_brutes),
        nombre_valides=len(lignes_valides),
        nombre_invalides=len(lignes_invalides),
        lignes_valides=lignes_valides,
        lignes_invalides=lignes_invalides,
    )


def generer_modele_excel() -> bytes:
    classeur = Workbook()
    feuille = classeur.active
    feuille.title = "Établissements"

    en_tete_style = Font(bold=True, color="FFFFFF")
    en_tete_fond = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

    for i, colonne in enumerate(COLONNES_ATTENDUES, start=1):
        cellule = feuille.cell(row=1, column=i, value=colonne)
        cellule.font = en_tete_style
        cellule.fill = en_tete_fond
        feuille.column_dimensions[get_column_letter(i)].width = 20

    feuille.append([
        "Clinique Exemple", "Littoral", "Cotonou", "1er Arrondissement", "Quartier Exemple",
        "Rue 123, quartier Exemple", "Clinique", "actif", "+22901XXXXXXXX", "6.3703", "2.3912",
    ])

    feuille_aide = classeur.create_sheet("Valeurs autorisées")
    feuille_aide.append(["Colonne", "Valeurs autorisées / format"])
    feuille_aide.append(["type", ", ".join(TYPES_ETABLISSEMENT_VALIDES)])
    feuille_aide.append(["statut", ", ".join(STATUTS_ETABLISSEMENT_VALIDES) + " (par défaut : actif)"])
    feuille_aide.append(["telephone", "+229 suivi de 01 puis 8 chiffres, ex : +22901XXXXXXXX"])
    feuille_aide.append(["departement / commune / arrondissement / quartier", "Doivent correspondre exactement au découpage territorial officiel (utilisez les mêmes intitulés que dans le formulaire de création manuelle)."])
    feuille_aide.append(["latitude / longitude", "Obligatoires — coordonnées GPS décimales, ex : 6.3703 / 2.3912"])
    feuille_aide.column_dimensions["A"].width = 55
    feuille_aide.column_dimensions["B"].width = 60

    tampon = BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()
