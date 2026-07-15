"""
Génération des exports PDF/Excel du rapport annuel — construits à partir de
la même structure de données que l'endpoint JSON /dashboard/rapports-mensuels
(voir app.routes.dashboard._construire_rapport_annuel), jamais de valeurs figées.
"""
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def _fmt_variation(valeur: int | float | None, suffixe: str = "%") -> str:
    if valeur is None:
        return "—"
    signe = "+" if valeur >= 0 else ""
    return f"{signe}{valeur}{suffixe}"


def _lignes_cumul(cumul: dict) -> list[tuple[str, str, str]]:
    return [
        ("Consultations (cumul annuel)", f"{cumul['consultations_ytd']:,}".replace(",", " "), _fmt_variation(cumul["consultations_ytd_variation"])),
        ("Patients actifs", f"{cumul['patients_actifs']:,}".replace(",", " "), _fmt_variation(cumul["patients_actifs_variation"])),
        ("Taux de couverture", f"{cumul['taux_couverture']}%", _fmt_variation(cumul["taux_couverture_variation_pts"], " pts")),
        ("Établissements actifs", f"{cumul['etablissements_actifs']}/{cumul['etablissements_total']}", "—"),
        ("Ordonnances émises", f"{cumul['ordonnances_emises']:,}".replace(",", " "), _fmt_variation(cumul["ordonnances_emises_variation"])),
        ("Alertes sécurité", str(cumul["alertes_securite"]), _fmt_variation(cumul["alertes_securite_variation"], "")),
    ]


def generer_pdf(rapport: dict) -> bytes:
    tampon = BytesIO()
    doc = SimpleDocTemplate(
        tampon, pagesize=A4,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    titre_style = ParagraphStyle("TitreDMPI", parent=styles["Title"], fontSize=18, spaceAfter=4)
    sous_titre_style = ParagraphStyle("SousTitre", parent=styles["Normal"], textColor=colors.grey, spaceAfter=16)
    section_style = ParagraphStyle("Section", parent=styles["Heading2"], spaceBefore=14, spaceAfter=8)
    sous_section_style = ParagraphStyle("SousSection", parent=styles["Heading3"], spaceBefore=10, spaceAfter=6)

    elements = []
    elements.append(Paragraph(f"Rapport annuel DMPI Bénin — {rapport['annee']}", titre_style))
    elements.append(Paragraph(f"Généré le {rapport['genere_le'][:10]} — réseau national des établissements de santé", sous_titre_style))

    elements.append(Paragraph("Indicateurs clés (cumul annuel)", section_style))
    donnees_cumul = [["Indicateur", "Valeur", "Variation vs N-1"]] + _lignes_cumul(rapport["cumul_annuel"])
    table_cumul = Table(donnees_cumul, colWidths=[8 * cm, 4 * cm, 4 * cm])
    table_cumul.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table_cumul)

    def _table_repartition(titre: str, entetes: list[str], lignes: list[list[str]], largeurs: list[float]) -> None:
        elements.append(Paragraph(titre, section_style))
        if not lignes:
            elements.append(Paragraph("Aucune donnée disponible pour l'année en cours.", styles["Normal"]))
            return
        table = Table([entetes] + lignes, colWidths=largeurs)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3A5F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(table)

    _table_repartition(
        "Répartition par département",
        ["Département", "Consultations", "Patients", "Établissements actifs"],
        [
            [d["departement"], str(d["consultations"]), str(d["patients"]), f"{d['etablissements_actifs']}/{d['etablissements_total']}"]
            for d in rapport["repartition_departements"]
        ],
        [6 * cm, 3.5 * cm, 3 * cm, 3.5 * cm],
    )

    _table_repartition(
        "Répartition par type d'établissement",
        ["Type", "Consultations", "Établissements"],
        [[t["type"], str(t["consultations"]), str(t["etablissements"])] for t in rapport["repartition_types_etablissement"]],
        [6 * cm, 5 * cm, 5 * cm],
    )

    _table_repartition(
        "Évolution mensuelle des consultations",
        ["Mois", "Consultations", "Patients"],
        [[m["mois"], str(m["consultations"]), str(m["patients"])] for m in reversed(rapport["rapports_mensuels"])],
        [6 * cm, 5 * cm, 5 * cm],
    )

    for mois in rapport["rapports_mensuels"]:
        elements.append(PageBreak())
        elements.append(Paragraph(f"Rapport mensuel — {mois['mois']}", section_style))

        donnees_chiffres = [
            ["Consultations", "Patients suivis", "Ordonnances", "Établissements actifs"],
            [str(mois["consultations"]), str(mois["patients"]), str(mois["ordonnances"]), str(mois["etablissements"])],
        ]
        table_chiffres = Table(donnees_chiffres, colWidths=[4 * cm] * 4)
        table_chiffres.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EDF3")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(table_chiffres)
        elements.append(Spacer(1, 0.4 * cm))
        elements.append(Paragraph(f"Taux de couverture : {mois['tauxCouverture']}%", styles["Normal"]))

        elements.append(Paragraph("Top diagnostics CIM-10", sous_section_style))
        if mois["topDiagnostics"]:
            donnees_diag = [["Code", "Libellé", "Cas"]] + [
                [d["code"], d["libelle"], str(d["count"])] for d in mois["topDiagnostics"]
            ]
            table_diag = Table(donnees_diag, colWidths=[2.5 * cm, 9 * cm, 2.5 * cm])
            table_diag.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EDF3")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            elements.append(table_diag)
        else:
            elements.append(Paragraph("Aucune consultation enregistrée ce mois-ci.", styles["Normal"]))

        elements.append(Paragraph("Top établissements par consultations", sous_section_style))
        if mois["topEtablissements"]:
            donnees_etabs = [["Établissement", "Consultations"]] + [
                [e["nom"], str(e["consultations"])] for e in mois["topEtablissements"]
            ]
            table_etabs = Table(donnees_etabs, colWidths=[10 * cm, 4 * cm])
            table_etabs.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EDF3")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            elements.append(table_etabs)
        else:
            elements.append(Paragraph("Aucune donnée d'établissement pour ce mois.", styles["Normal"]))

    doc.build(elements)
    return tampon.getvalue()


def generer_excel(rapport: dict) -> bytes:
    classeur = Workbook()
    entete_police = Font(bold=True, color="FFFFFF")
    entete_fond = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")

    def _entete(feuille, colonnes: list[str]) -> None:
        feuille.append(colonnes)
        for cellule in feuille[1]:
            cellule.font = entete_police
            cellule.fill = entete_fond

    feuille_cumul = classeur.active
    feuille_cumul.title = "Cumul annuel"
    _entete(feuille_cumul, ["Indicateur", "Valeur", "Variation vs N-1"])
    for ligne in _lignes_cumul(rapport["cumul_annuel"]):
        feuille_cumul.append(list(ligne))
    feuille_cumul.column_dimensions["A"].width = 32
    feuille_cumul.column_dimensions["B"].width = 16
    feuille_cumul.column_dimensions["C"].width = 18

    feuille_mensuel = classeur.create_sheet("Rapport mensuel")
    _entete(feuille_mensuel, ["Mois", "Consultations", "Patients", "Ordonnances", "Établissements actifs", "Taux de couverture (%)"])
    for m in rapport["rapports_mensuels"]:
        feuille_mensuel.append([m["mois"], m["consultations"], m["patients"], m["ordonnances"], m["etablissements"], m["tauxCouverture"]])
    for col, largeur in zip("ABCDEF", [18, 14, 12, 14, 20, 20]):
        feuille_mensuel.column_dimensions[col].width = largeur

    feuille_diag = classeur.create_sheet("Top diagnostics")
    _entete(feuille_diag, ["Mois", "Code CIM-10", "Libellé", "Cas"])
    for m in rapport["rapports_mensuels"]:
        for d in m["topDiagnostics"]:
            feuille_diag.append([m["mois"], d["code"], d["libelle"], d["count"]])
    for col, largeur in zip("ABCD", [18, 14, 45, 10]):
        feuille_diag.column_dimensions[col].width = largeur

    feuille_etabs = classeur.create_sheet("Top établissements")
    _entete(feuille_etabs, ["Mois", "Établissement", "Consultations"])
    for m in rapport["rapports_mensuels"]:
        for e in m["topEtablissements"]:
            feuille_etabs.append([m["mois"], e["nom"], e["consultations"]])
    for col, largeur in zip("ABC", [18, 35, 14]):
        feuille_etabs.column_dimensions[col].width = largeur

    feuille_departements = classeur.create_sheet("Répartition départements")
    _entete(feuille_departements, ["Département", "Consultations", "Patients", "Établissements actifs", "Établissements total"])
    for d in rapport["repartition_departements"]:
        feuille_departements.append([d["departement"], d["consultations"], d["patients"], d["etablissements_actifs"], d["etablissements_total"]])
    for col, largeur in zip("ABCDE", [22, 14, 12, 18, 18]):
        feuille_departements.column_dimensions[col].width = largeur

    feuille_types = classeur.create_sheet("Répartition types établissement")
    _entete(feuille_types, ["Type d'établissement", "Consultations", "Établissements"])
    for t in rapport["repartition_types_etablissement"]:
        feuille_types.append([t["type"], t["consultations"], t["etablissements"]])
    for col, largeur in zip("ABC", [22, 14, 14]):
        feuille_types.column_dimensions[col].width = largeur

    tampon = BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()
